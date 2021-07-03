import time
import datetime, openpyxl, sys
from openpyxl.styles import borders, Font
from openpyxl.styles.fills import PatternFill
from bs4 import BeautifulSoup
from urllib.request import Request
import urllib

slownik = {}
aktualnaData = str(datetime.datetime.today().strftime('%Y-%m-%d'))
#aktualnaData = '2021-03-15'


def pobieranieMorele():
    linki = ['https://www.morele.net/procesor-amd-ryzen-5-3600-3-6ghz-32-mb-box-100-100000031box-5938599/',
             '',
             '',
             # 'https://www.morele.net/karta-graficzna-sapphire-radeon-rx-5700-xt-pulse-8gb-gddr6-11293-01-20g-6276799/',
             '',
             '',  # 'https://www.morele.net/plyta-glowna-asus-tuf-b450-pro-gaming-5566520/',
             'https://www.morele.net/plyta-glowna-msi-mpg-x570-gaming-edge-wifi-5938606/',
             '',
             'https://www.morele.net/dysk-ssd-adata-xpg-gammix-s11-pro-512-gb-m-2-2280-pci-e-x4-gen3-nvme-agammixs11p-512gt-c-5625912/',
             '',
             '',
             'https://www.morele.net/pamiec-goodram-irdm-ddr4-16-gb-3600mhz-cl17-irp-3600d4v64l17s-16gdc-6432278/',
             'https://www.morele.net/pamiec-crucial-ballistix-rgb-black-at-ddr4-3600-16gb-cl16-bl2k8g36c16u4bl-6492649/',
             '',
             'https://www.morele.net/zasilacz-silentiumpc-supremo-m2-gold-550w-spc140-774137/',
             'https://www.morele.net/zasilacz-silentiumpc-supremo-fm2-gold-650w-spc168-1243754/',
             'https://www.morele.net/chlodzenie-cpu-silentiumpc-fortis-3-rgb-spc245-5940755/',
             'https://www.morele.net/chlodzenie-cpu-silentiumpc-fortis-3-evo-argb-he1425-spc278-6575309/',
             '',
             'https://www.morele.net/obudowa-silentiumpc-armis-ar7x-evo-tg-argb-spc251-6535607/',
             'https://www.morele.net/obudowa-silentiumpc-armis-ar6q-evo-tg-argb-ze-szklanym-oknem-spc256-5941407/',
             'https://www.morele.net/obudowa-silentiumpc-signum-sg1q-evo-tg-argb-pure-black-spc253-6524283/',
             'https://www.morele.net/obudowa-silentiumpc-regnum-rg6v-tg-spc261-5941313/',
             'https://www.morele.net/obudowa-silentiumpc-regnum-rg6v-evo-tg-argb-spc262-5941314/',
             'https://www.morele.net/obudowa-msi-mag-vampiric-010-4145491/',
             '',
             'https://www.morele.net/monitor-aoc-agon-ag241qx-1060690/',
             '',
             'https://www.morele.net/chlodzenie-cpu-silentiumpc-grandis-3-spc274-5942593/',
             'https://www.morele.net/plyta-glowna-asus-rog-strix-b550-e-gaming-90mb1470-m0eay0-6858395/',
             'https://www.morele.net/plyta-glowna-asus-rog-strix-b550-f-gaming-wi-fi-6787910/',
             'https://www.morele.net/dysk-ssd-patriot-p210-1-tb-2-5-sata-iii-p210s1tb25-5942962/',
             'https://www.morele.net/pamiec-patriot-viper-4-blackout-ddr4-16-gb-3600mhz-cl18-pvb416g360c8k-6957974/',
             '',
             'https://www.morele.net/karta-graficzna-asus-tuf-radeon-rx-6700-xt-gaming-oc-12gb-gddr6-tuf-rx6700xt-o12g-gaming-5946539/',
             'https://www.morele.net/karta-graficzna-power-color-radeon-rx-6700-xt-12gb-gddr6-axrx-6700xt-12gbd6-m3dh-5946848/',
             'https://www.morele.net/karta-graficzna-power-color-radeon-rx-6700-xt-red-devil-12gb-gddr6-axrx-6700xt-12gbd6-3dhe-oc-5946847/',
             'https://www.morele.net/karta-graficzna-xfx-radeon-rx-6700-xt-speedtester-qick319-ultra-12gb-gddr6-rx-67xtypudp-8148720/',
             'https://www.morele.net/karta-graficzna-msi-radeon-rx-6700-xt-mech-2x-oc-12gb-gddr6-rx-6700-xt-mech-2x-12g-oc-5946757/',
             'https://www.morele.net/karta-graficzna-xfx-radeon-rx-6700-xt-speedtester-merc319-black-12gb-gddr6-rx-67xtytbdp-8148721/',
             'https://www.morele.net/karta-graficzna-msi-radeon-rx-6700-xt-gaming-x-12gb-gddr6-rx-6700-xt-gaming-x-12g-5946756/',
             'https://www.morele.net/karta-graficzna-sapphire-radeon-rx-6700-xt-pulse-gaming-12gb-gddr6-11306-02-20g-5946772/',
             'https://www.morele.net/karta-graficzna-sapphire-radeon-rx-6700-xt-nitro-gaming-oc-11306-01-20g-5946773/',
             'https://www.morele.net/karta-graficzna-asus-radeon-rx-6700-xt-12gb-gddr6-rx6700xt-12g-5946805/',
             'https://www.morele.net/karta-graficzna-gigabyte-radeon-rx-6700-xt-gaming-oc-12gb-gddr6-gv-r67xtgaming-oc-12gd-5946529/',
             'https://www.morele.net/karta-graficzna-asus-rog-strix-radeon-rx-6700-xt-gaming-oc-12gb-gddr6-rog-strix-rx6700xt-o12g-gaming-5946540/',
             'https://www.morele.net/karta-graficzna-asrock-radeon-rx-6700-xt-12gb-gddr6-90-ga2wzz-00uanf-5946712/',
             'https://www.morele.net/karta-graficzna-msi-radeon-rx-6700-xt-12gb-gddr6-rx-6700-xt-12g-5946755/',
             '',
             '',
             '',
             '',
             '',
             'https://www.morele.net/karta-graficzna-gigabyte-gigabyte-radeon-rx-6800-gaming-oc-16g-gv-r68gaming-oc-16gd-7497486/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-radeon-rx-6800-master-16gb-gddr6-gv-r68aorus-m-16gd-7497485/',
             'https://www.morele.net/karta-graficzna-asrock-radeon-rx-6800-xt-phantom-gaming-d-oc-16gb-gddr6-rx6800xt-pgd-16go-7497491/',
             'https://www.morele.net/karta-graficzna-msi-rx-6800-xt-gaming-x-trio-16gb-gddr6-rx-6800-xt-gaming-x-trio-5945763/',
             'https://www.morele.net/karta-graficzna-power-color-radeon-6800xt-red-devil-16gb-gddr6-axrx-6800xt-16gbd6-3dhe-oc-5946036/',
             'https://www.morele.net/karta-graficzna-sapphire-radeon-rx-6800-xt-pulse-16gb-gddr6-11304-03-20g-5947823/',
             'https://www.morele.net/karta-graficzna-asus-tuf-radeon-rx-6800-gaming-oc-16gb-gddr6-tuf-rx6800-o16g-gaming-7497493/',
             '',
             '',
             '',
             '',
             '',
             'https://www.morele.net/karta-graficzna-msi-radeon-rx-6900-xt-gaming-x-trio-16gb-gddr6-rx-6900-xt-gaming-x-trio-5946901/',
             'https://www.morele.net/karta-graficzna-sapphire-radeon-rx-6900-xt-nitro-se-oc-16gb-gddr6-11308-03-20g-8260735/',
             'https://www.morele.net/karta-graficzna-asrock-radeon-rx-6900-xt-oc-formula-16gb-gddr6-rx6900xt-ocf-16g-5947341/',
             'https://www.morele.net/karta-graficzna-msi-radeon-rx-6900-xt-gaming-z-trio-16gb-gddr6-rx-6900-xt-gaming-z-trio-5947791/',
             'https://www.morele.net/karta-graficzna-power-color-radeon-rx-6900xt-ultimate-16gb-gddr6-axrx-6900xtu-16gbd6-3dhe-oc-5947694/',
             'https://www.morele.net/karta-graficzna-sapphire-toxic-radeon-rx-6900-xt-gaming-oc-16gb-gddr6-11308-06-20g-5948002/',
             'https://www.morele.net/karta-graficzna-sapphire-toxic-radeon-rx-6900-xt-gaming-oc-extreme-edition-11308-08-20g-5947696/',
             'https://www.morele.net/karta-graficzna-gigabyte-radeon-rx-6900-xt-gaming-oc-16gb-gddr6-gv-r69xtgaming-oc-16gd-5945953/',
             'https://www.morele.net/karta-graficzna-asrock-radeon-rx-6900-xt-phantom-gaming-d-oc-16gb-gddr6-rx6900xt-pgd-16go-5945915/',
             'https://www.morele.net/karta-graficzna-asus-tuf-radeon-rx-6900-xt-gaming-oc-16gb-gddr6-tuf-rx6900xt-o16g-gaming-5945886/',
             'https://www.morele.net/karta-graficzna-power-color-radeon-rx-6900xt-red-devil-16gb-gddr6-axrx-6900xt-16gbd6-3dhe-oc-5947999/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-radeon-rx-6900-xt-master-16gb-gddr6-gv-r69xtaorus-m-16gd-5946682/',
             '',
             '',
             '',
             '',
             '',
             'https://www.morele.net/karta-graficzna-kfa2-geforce-rtx-3060-1-click-oc-12gb-gddr6-36nol7md1vok-5946547/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3060-gaming-oc-12g-gv-n3060gaming-oc-12gd-7912037/',
             'https://www.morele.net/karta-graficzna-inno3d-geforce-rtx-3060-twin-x2-oc-12gb-gddr6-n30602-12d6x-11902120-5946320/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3060-gaming-x-trio-12gb-gddr6-rtx-3060-gaming-x-trio-12g-5946237/',
             'https://www.morele.net/karta-graficzna-asus-rog-strix-geforce-rtx-3060-gaming-oc-12gb-gddr6-rog-strix-rtx3060-o12ggaming-7951868/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3060-eagle-12gb-gddr6-gv-n3060eagle-12gd-7912040/',
             'https://www.morele.net/karta-graficzna-asus-dual-geforce-rtx-3060-gaming-oc-12gb-gddr6-dual-rtx3060-o12g-v2-5948085/',
             'https://www.morele.net/karta-graficzna-asus-phoenix-geforce-rtx-3060-12gb-gddr6-ph-rtx3060-12g-5947627/',
             'https://www.morele.net/karta-graficzna-asus-tuf-geforce-rtx-3060-gaming-oc-12gb-gddr6-tuf-rtx3060-o12g-v2-gaming-5948083/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3060-ventus-x2-oc-12gb-gddr6-rtx-3060-ventus-2x-12g-oc-5946240/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3060-vision-oc-12gb-gddr6-gv-n3060vision-oc-12gd-7912038/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3060-elite-12gb-gddr6-gv-n3060aorus-e-12gd-7912036/',
             'https://www.morele.net/karta-graficzna-asus-tuf-geforce-rtx-3060-gaming-oc-12gb-gddr6-tuf-rtx3060-o12g-gaming-7951867/',
             'https://www.morele.net/karta-graficzna-zotac-geforce-rtx-3060-amp-white-edition-12gb-gddr6-zt-a30600f-10p-7951869/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3060-elite-12gb-gddr6-gv-n3060aorus-e-12gd-2-0-5948079/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3060-ventus-x3-oc-12gb-gddr6-rtx-3060-ventus-3x-12g-oc-5946239/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3060-vision-oc-12gb-gddr6-gv-n3060vision-oc-12gd-2-0-5948082/',
             'https://www.morele.net/karta-graficzna-palit-geforce-rtx-3060-stormx-oc-12gb-gddr6-ne63060s19k9-190af-5946396/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3060-gaming-oc-12g-gv-n3060gaming-oc-12gd-2-0-5948081/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3060-eagle-oc-12gb-gddr6-gv-n3060eagle-oc-12gd-2-0-5948080/',
             '',
             '',
             '',
             '',
             'https://www.morele.net/karta-graficzna-asus-dual-geforce-rtx-3070-gaming-8gb-gddr6-dual-rtx3070-8g-7244530/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3070-gaming-x-trio-8gb-gddr6-rtx-3070-gaming-x-trio-7244535/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3070-gaming-oc-8gb-gddr6-gv-n3070gaming-oc-8gd-5944660/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3070-master-8gb-gddr6-gv-n3070aorus-m-8gd-7244534/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3070-ventus-3x-oc-8gb-gddr6-rtx-3070-ventus-3x-oc-7244536/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3070-vision-oc-8gb-gddr6-gv-n3070vision-oc-8gd-7244533/',
             'https://www.morele.net/karta-graficzna-palit-geforce-rtx-3070-gamingpro-8gb-gddr6-ne63070019p2-1041a-5944665/',
             'https://www.morele.net/karta-graficzna-inno3d-geforce-rtx-3070-ichill-x4-8gb-gddr6-rtx-3070-ichill-x4-5943767/',
             'https://www.morele.net/karta-graficzna-palit-geforce-rtx-3070-jetstream-8gb-gddr6-ne63070019p2-1040j-5945386/',
             'https://www.morele.net/karta-graficzna-inno3d-geforce-rtx-3070-twin-x2-oc-8gb-gddr6-rtx-3070-twin-x2-oc-7698848/',
             'https://www.morele.net/karta-graficzna-gainward-geforce-rtx-3070-phantom-8gb-gddr6-471056224-2171-7698846/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3070-suprim-x-8gb-gddr6-rtx-3070-suprim-x-8g-5945714/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3070-ventus-2x-oc-8gb-gddr6-rtx-3070-ventus-2x-oc-7244537/',
             'https://www.morele.net/karta-graficzna-asus-tuf-geforce-rtx-3070-gaming-oc-8gb-gddr6-tuf-rtx3070-o8g-gaming-7244531/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3070-eagle-oc-8gb-gddr6-gv-n3070eagle-oc-8gd-5944661/',
             'https://www.morele.net/karta-graficzna-zotac-geforce-rtx-3070-twin-edge-oc-8gb-gddr6-zt-a30700h-10p-7463163/',
             'https://www.morele.net/karta-graficzna-kfa2-geforce-rtx-3070-8gb-gddr6-37nsl6md2kok-5946073/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3070-gaming-z-trio-8gb-gddr6-rtx-3070-gaming-z-trio-5947828/',
             '',
             '',
             '',
             '',
             '',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3070-ti-suprim-x-8gb-gddr6x-rtx-3070ti-suprim-x-8g-5947971/',
             'https://www.morele.net/karta-graficzna-inno3d-geforce-rtx-3070-ti-x3-oc-dual-slot-8gb-gddr6x-n307t3-086xx-1820va45-5947969/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3070-ti-master-8gb-gddr6x-gv-n307taorus-m-8gd-5948092/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3070-ti-gaming-x-trio-8gb-gddr6x-rtx-3070-ti-gaming-x-trio-8g-5948060/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3070-ti-ventus-3x-oc-8gb-gddr6x-rtx-3070-ti-ventus-3x-5947972/',
             'https://www.morele.net/karta-graficzna-palit-geforce-rtx-3070-ti-gamingpro-oc-8gb-gddr6x-ned307t019p2-1046a-5948006/',
             'https://www.morele.net/karta-graficzna-gainward-geforce-rtx-3070-ti-phoenix-8gb-gddr6x-471056224-2713-5948007/',
             'https://www.morele.net/karta-graficzna-palit-geforce-rtx-3070-ti-gamerock-oc-8gb-gddr6x-ned307tt19p2-1047g-5948005/',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3080-gaming-oc-10gb-gddr6x-gv-n3080gaming-oc-10gd-5943775/',
             'https://www.morele.net/karta-graficzna-inno3d-geforce-rtx-3080-ichill-x4-10gb-gddr6x-c30804-106xx-1810va36-5943908/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3080-suprim-x-10gb-gddr6x-rtx-3080-suprim-x-10g-5945417/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3080-xtreme-waterforce-wb-10gb-gddr6x-gv-n3080aorusx-wb-10gd-5945408/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3080-turbo-10gb-gddr6x-gv-n3080turbo-10gd-5946829/',
             'https://www.morele.net/karta-graficzna-zotac-geforce-rtx-3080-trinity-oc-10gb-gddr6x-zt-a30800j-10p-5946596/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3080-gaming-z-trio-10gb-gddr6-rtx-3080-gaming-z-trio-10g-5947089/',
             'https://www.morele.net/karta-graficzna-kfa2-geforce-rtx-3080-sg-10gb-gddr6x-38nwm3md99nk-5947897/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3080-gaming-oc-waterforce-10gb-gddr6x-gv-n3080gamingoc-wb-10gd-5947455/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3080-xtreme-10gb-gddr6x-gv-n3080aorus-x-10gd-5944923/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3080-xtreme-waterforce-10gb-gddr6x-gv-n3080aorusx-w-10gd-5945519/',
             'https://www.morele.net/karta-graficzna-asus-rog-strix-geforce-rtx-3080-gaming-oc-white-10gb-gddr6x-rog-strix-rtx3080-o10g-white-5945885/',
             'https://www.morele.net/karta-graficzna-zotac-geforce-rtx-3080-amp-holo-10gb-gddr6x-zt-a30800f-10p-5945574/',
             '',
             '',
             '',
             '',
             '',
             '',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3080-ti-gaming-oc-12gb-gddr6x-gv-n308tgaming-oc-12gd-5947950/',
             'https://www.morele.net/karta-graficzna-palit-geforce-rtx-3080-ti-gamerock-oc-12gb-gddr6x-ned308tt19kb-1020g-5948004/',
             'https://www.morele.net/karta-graficzna-palit-geforce-rtx-3080-ti-gamerock-12gb-gddr6x-ned308t019kb-1020g-5947952/',
             'https://www.morele.net/karta-graficzna-palit-geforce-rtx-3080-ti-gamingpro-12gb-gddr6x-ned308t019kb-132aa-5947954/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3080-ti-vision-oc-12gb-gddr6x-gv-n308tvision-oc-12gd-1-0-5947949/',
             'https://www.morele.net/karta-graficzna-gainward-geforce-rtx-3080-ti-phoenix-12gb-gddr6x-471056224-2379-5947953/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3080-ti-master-12gb-gddr6x-gv-n308taorus-m-12gd-5948094/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3080-ti-gaming-x-trio-12gb-gddr6x-rtx-3080ti-gaming-x-trio-5947955/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3080-ti-suprim-x-12gb-gddr6x-rtx-3080-ti-suprim-x-12g-5947956/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3080-ti-eagle-oc-12gb-gddr6x-gv-n308teagle-12gd-5947951/',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3090-gaming-x-trio-24gb-gddr6x-rtx-3090-gaming-x-trio-24g-5943773/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3090-suprim-x-24gb-gddr6x-5945166/',
             'https://www.morele.net/karta-graficzna-inno3d-geforce-rtx-3090-ichill-x3-24gb-gddr6x-rtx-3090-ichill-x3-7698847/',
             'https://www.morele.net/karta-graficzna-gigabyte-geforce-rtx-3090-gaming-oc-24gb-gddr6x-gv-n3090gaming-oc-24gd-5943778/',
             'https://www.morele.net/karta-graficzna-zotac-geforce-rtx-3090-trinity-oc-24gb-gddr6x-zt-a30900j-10p-5946556/',
             'https://www.morele.net/karta-graficzna-gigabyte-aorus-geforce-rtx-3090-xtreme-waterforce-24gb-gddr6x-gv-n3090aorusx-w-24gd-5945813/',
             'https://www.morele.net/karta-graficzna-kfa2-geforce-rtx-3090-hof-limited-edition-24gb-gddr6x-39nxm5md3blk-5947326/',
             'https://www.morele.net/karta-graficzna-kfa2-geforce-rtx-3090-sg-24gb-gddr6x-39nsm5md1gnk-5947179/',
             'https://www.morele.net/karta-graficzna-inno3d-geforce-rtx-3090-ichill-x4-24gb-gddr6x-rtx-3090-ichill-x4-5943761/',
             'https://www.morele.net/karta-graficzna-msi-geforce-rtx-3090-ventus-oc-3x-24gb-gddr6x-rtx-3090-ventus-3x-24g-oc-5943774/',
             'https://www.morele.net/karta-graficzna-pny-geforce-rtx-3090-xlr8-gaming-revel-edition-24gb-gddr6x-vcg309024tfxppb-5945482/',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '',
             '']
    i = 0
    for x in linki:
        try:
            if x != '':
                print(x)
                req = Request(x, headers={"User-Agent": "Mozilla/5.0"})
                response = urllib.request.urlopen(req)
                html = response.read()
                if response.getcode() != 200:
                    print('continue')
                    continue
                soup = BeautifulSoup(html, 'html.parser')
                slownik[i, 0] = soup.find('h1', {'class': ['prod-name']}).text  # nazwa
                slownik[i, 1] = soup.find('div', {'class': ['product-price']}).text  # aktualnaCena
                slownik[i, 1] = slownik[i, 1].replace("z", "")
                slownik[i, 1] = slownik[i, 1].replace("ł", "")
                slownik[i, 1] = slownik[i, 1].replace(" ", "")
                slownik[i, 2] = soup.find('div', {'class': ['product-price-old']})  # poprzedniaCena
                if slownik[i, 2]:
                    slownik[i, 2] = True
                else:
                    slownik[i, 2] = False
                slownik[i, 3] = soup.find('button',
                                          {'class': [
                                              'add-to-cart__disabled btn btn-grey btn-block btn-sidebar btn-disabled']})  # dostepnosc
                slownik[i, 3] = str(slownik[i, 3])
                if 'disabled' in slownik[i, 3] or 'NIEDOSTĘPNY' in slownik[i, 3]:
                    slownik[i, 3] = False
                else:
                    slownik[i, 3] = True
                slownik[i, 4] = x  # link
            else:
                slownik[i, 0] = ''
                slownik[i, 1] = ''
                slownik[i, 2] = ''
                slownik[i, 3] = ''
                slownik[i, 4] = ''
            i = i + 1
        except:
            i = i + 1
    return i - 1


def zapisDatyMorele(p):
    plik = openpyxl.load_workbook(p)
    odl = 0
    for x in range(5, 1000):
        if plik["Arkusz1"].cell(row=1, column=x).value == aktualnaData and plik["Arkusz1"].cell(row=2, column=x).value != None:
            print(plik["Arkusz1"].cell(row=2, column=x).value)
            return odl
        elif plik["Arkusz1"].cell(row=1, column=x).value is None or (plik["Arkusz1"].cell(row=1, column=x).value == aktualnaData and plik["Arkusz1"].cell(row=2, column=x).value == None):
            plik["Arkusz1"].cell(row=1, column=x).value = aktualnaData
            plik["Arkusz1"].cell(row=1, column=x).fill = PatternFill(fgColor="C6E0B4", fill_type="solid")
            odl = x
            plik.save(p)
            plik.close()
            return odl
    plik.save(p)
    plik.close()
    return odl


def zapisDanychMorele(odl, p):
    plik = openpyxl.load_workbook(p)
    i = 0
    ile = 0
    for x in slownik:
        # print(f'{x}  {ile}  {i}')
        # print(ile)
        if ile == 0:
            plik["Arkusz1"].cell(row=i + 2, column=2).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=2).value = slownik[i, 0]
            ile += 1
        elif ile == 1:
            if str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) > slownik[i, 1] and str(
                    plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) != None and slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="E6B8B7", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) > slownik[i, 1] and str(
                    plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) != None:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="FFE699", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) < slownik[i, 1] and slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="632523", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) < slownik[i, 1]:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="F4B084", fill_type="solid")
            elif slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=odl).value = slownik[i, 1]
            ile += 1
        elif ile == 2:
            if slownik[i, 2]:
                plik["Arkusz1"].cell(row=i + 2, column=4).fill = PatternFill(fgColor="00B050", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=4).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=4).value = slownik[i, 2]
            ile += 1
        elif ile == 3:
            if slownik[i, 3]:
                plik["Arkusz1"].cell(row=i + 2, column=3).fill = PatternFill(fgColor="00B050", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=3).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=3).value = slownik[i, 3]
            ile += 1

        else:
            plik["Arkusz1"].cell(row=i + 2, column=1).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=1).value = slownik[i, 4]
            ile = 0
            i += 1

    plik.save(p)
    plik.close()


def pobieranieXkom():
    linki = ['https://www.x-kom.pl/p/500085-procesor-amd-ryzen-5-amd-ryzen-5-3600.html',
             '',
             'https://www.x-kom.pl/p/521421-karta-graficzna-amd-sapphire-radeon-rx-5700-xt-pulse-8gb-gddr6.html',
             '',
             'https://www.x-kom.pl/p/493301-plyta-glowna-socket-am4-asus-tuf-b450-pro-gaming.html',
             'https://www.x-kom.pl/p/500398-plyta-glowna-socket-am4-msi-mpg-x570-gaming-edge-wifi.html',
             '',
             'https://www.x-kom.pl/p/474474-dysk-ssd-adata-512gb-m2-pcie-nvme-xpg-gammix-s11-pro.html',
             '',
             '',
             'https://www.x-kom.pl/p/531223-pamiec-ram-ddr4-goodram-16gb-2x8gb-3600mhz-cl17-irdm-pro.html',
             'https://www.x-kom.pl/p/550277-pamiec-ram-ddr4-crucial-16gb-2x8gb-3600mhz-cl16-ballistix-black-rgb.html',
             '',
             'https://www.x-kom.pl/p/308096-zasilacz-do-komputera-silentiumpc-supremo-m2-550w-80-plus-gold.html',
             'https://www.x-kom.pl/p/363851-zasilacz-do-komputera-silentiumpc-supremo-fm2-650w-80-plus-gold.html',
             'https://www.x-kom.pl/p/529353-chlodzenie-procesora-silentiumpc-fortis-3-rgb-140mm.html',
             'https://www.x-kom.pl/p/550449-chlodzenie-procesora-silentiumpc-fortis-3-evo-argb-140mm.html',
             '',
             'https://www.x-kom.pl/p/546569-obudowa-do-komputera-silentiumpc-armis-ar7x-evo-tg-argb.html',
             'https://www.x-kom.pl/p/544983-obudowa-do-komputera-silentiumpc-armis-ar6q-evo-tg-argb.html',
             'https://www.x-kom.pl/p/548256-obudowa-do-komputera-silentiumpc-signum-sg1q-evo-tg-argb.html',
             'https://www.x-kom.pl/p/541990-obudowa-do-komputera-silentiumpc-regnum-rg6v-tg-pure-black.html',
             'https://www.x-kom.pl/p/542017-obudowa-do-komputera-silentiumpc-regnum-rg6v-evo-tg-argb.html',
             'https://www.x-kom.pl/p/491808-obudowa-do-komputera-msi-mag-vampiric-010.html',
             '',
             'https://www.x-kom.pl/p/333314-monitor-led-24-aoc-agon-ag241qx.html',
             'https://www.x-kom.pl/p/359197-sluchawki-bezprzewodowe-steelseries-arctis-7-czarne-bezprzewodowe.html',
             '',
             'https://www.x-kom.pl/p/566501-chlodzenie-procesora-silentiumpc-grandis-3-120-140mm.html',
             'https://www.x-kom.pl/p/569335-plyta-glowna-socket-am4-asus-rog-strix-b550-f-gaming-wi-fi.html',
             'https://www.x-kom.pl/p/575331-dysk-ssd-patriot-1tb-25-sata-ssd-p210.html',
             'https://www.x-kom.pl/p/591878-pamiec-ram-ddr4-patriot-16gb-2x8gb-3600mhz-cl18-viper-steel.html',
             'https://www.x-kom.pl/p/591220-karta-graficzna-amd-sapphire-radeon-rx-5700-xt-pulse-be-8gb-gddr6.html',
             'https://www.x-kom.pl/p/515067-karta-graficzna-amd-powercolor-radeon-rx-5700-xt-red-dragon-8gb-gddr6.html',
             '',
             '',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/609100-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-ti-gaming-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/609097-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-ti-aorus-master-8gb-gddr6.html',
             'https://www.x-kom.pl/p/608934-karta-graficzna-nvidia-msi-geforce-rtx-3060-ti-gaming-x-trio-8gb-gddr6.html',
             'https://www.x-kom.pl/p/626062-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-ti-vision-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/609936-karta-graficzna-nvidia-asus-geforce-rtx-3060-ti-tuf-gaming-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/625228-karta-graficzna-nvidia-msi-geforce-rtx-3060-ti-gaming-x-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610301-karta-graficzna-nvidia-zotac-geforce-rtx-3060-ti-twin-edge-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/609939-karta-graficzna-nvidia-asus-geforce-rtx-3060-ti-rog-strix-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610722-karta-graficzna-nvidia-gainward-geforce-rtx-3060-ti-phoenix-gs-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610297-karta-graficzna-nvidia-palit-geforce-rtx-3060-ti-dual-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610720-karta-graficzna-nvidia-palit-geforce-rtx-3060-ti-gaming-pro-8gb-gddr6.html',
             'https://www.x-kom.pl/p/609943-karta-graficzna-nvidia-asus-geforce-rtx-3060-ti-dual-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610296-karta-graficzna-nvidia-palit-geforce-rtx-3060-ti-dual-8gb-gddr6.html',
             'https://www.x-kom.pl/p/609099-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-ti-gaming-oc-pro-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610721-karta-graficzna-nvidia-gainward-geforce-rtx-3060-ti-phoenix-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610718-karta-graficzna-nvidia-palit-geforce-rtx-3060-ti-gaming-pro-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610299-karta-graficzna-nvidia-zotac-geforce-rtx-3060-ti-twin-edge-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610294-karta-graficzna-nvidia-inno3d-geforce-rtx-3060-ti-ichill-x3-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610293-karta-graficzna-nvidia-inno3d-geforce-rtx-3060-ti-twin-x2-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/610290-karta-graficzna-nvidia-gainward-geforce-rtx-3060-ti-ghost-oc-8gb.html',
             'https://www.x-kom.pl/p/610289-karta-graficzna-nvidia-gainward-geforce-rtx-3060-ti-ghost-8gb-gddr6.html',
             'https://www.x-kom.pl/p/609102-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-ti-eagle-8gb-gddr6.html',
             'https://www.x-kom.pl/p/608936-karta-graficzna-nvidia-msi-geforce-rtx-3060-ti-ventus-3x-oc-8gb-gddr6.html',
             '',
             '',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/597348-karta-graficzna-nvidia-msi-geforce-rtx-3070-gaming-x-trio-8gb-gddr6.html',
             'https://www.x-kom.pl/p/620469-karta-graficzna-nvidia-zotac-geforce-rtx-3070-twin-edge-oc-white-8gb-gddr6.html',
             'https://www.x-kom.pl/p/597350-karta-graficzna-nvidia-msi-geforce-rtx-3070-ventus-3x-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/607561-karta-graficzna-nvidia-zotac-geforce-rtx-3070-twin-edge-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/619845-karta-graficzna-nvidia-palit-geforce-rtx-3070-jetstream-8gb-gddr6.html',
             'https://www.x-kom.pl/p/623536-karta-graficzna-nvidia-gainward-geforce-rtx-3070-phantom-gs-8gb-gddr6.html',
             'https://www.x-kom.pl/p/602344-karta-graficzna-nvidia-gainward-geforce-rtx-3070-phoenix-gs-8gb-gddr6.html',
             'https://www.x-kom.pl/p/596768-karta-graficzna-nvidia-asus-geforce-rtx-3070-rog-strix-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/607116-karta-graficzna-nvidia-gainward-geforce-rtx-3070-phoenix-8gb-gddr6.html',
             'https://www.x-kom.pl/p/604786-karta-graficzna-nvidia-asus-geforce-rtx-3070-dual-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/622493-karta-graficzna-nvidia-asus-geforce-rtx-3070-rog-strix-oc-white-8gb-gddr6.html',
             'https://www.x-kom.pl/p/602341-karta-graficzna-nvidia-palit-geforce-rtx-3070-gaming-pro-8gb-gddr6.html',
             'https://www.x-kom.pl/p/622036-karta-graficzna-nvidia-palit-geforce-rtx-3070-gamerock-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/612835-karta-graficzna-nvidia-palit-geforce-rtx-3070-gamerock-8gb-gddr6.html',
             'https://www.x-kom.pl/p/627753-karta-graficzna-nvidia-inno3d-geforce-rtx-3070-ichill-x3-8gb-gddr6.html',
             'https://www.x-kom.pl/p/622031-karta-graficzna-nvidia-palit-geforce-rtx-3070-jetstream-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/589763-karta-graficzna-nvidia-zotac-geforce-rtx-3070-gaming-twin-edge-8gb-gddr6.html',
             'https://www.x-kom.pl/p/622043-karta-graficzna-nvidia-inno3d-geforce-rtx-3070-twin-x2-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/604606-karta-graficzna-nvidia-msi-geforce-rtx-3070-suprim-x-8gb-gddr6.html',
             'https://www.x-kom.pl/p/602535-karta-graficzna-nvidia-asus-geforce-rtx-3070-dual-8gb-gddr6.html',
             'https://www.x-kom.pl/p/602343-karta-graficzna-nvidia-palit-geforce-rtx-3070-gaming-pro-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/597344-karta-graficzna-nvidia-gigabyte-geforce-rtx-3070-eagle-8gb-gddr6.html',
             '',
             '',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/589756-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-gaming-oc-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/589740-karta-graficzna-nvidia-msi-geforce-rtx-3080-gaming-x-trio-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/590074-karta-graficzna-nvidia-asus-geforce-rtx-3080-tuf-gaming-oc-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/600904-karta-graficzna-nvidia-msi-geforce-rtx-3080-suprim-x-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/607813-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-xtreme-waterforce-wb-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/589742-karta-graficzna-nvidia-msi-geforce-rtx-3080-ventus-3x-oc-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/622492-karta-graficzna-nvidia-asus-geforce-rtx-3080-rog-strix-oc-white-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/613194-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-aorus-xtreme-wf-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/604635-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-aorus-xtreme-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/596928-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-aorus-master-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/592545-karta-graficzna-nvidia-asus-geforce-rtx-3080-rog-strix-oc-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/625242-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-aorus-master-10gb-gddr6x-rev-20.html',
             'https://www.x-kom.pl/p/613189-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-eagle-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/590450-karta-graficzna-nvidia-palit-geforce-rtx-3080-gaming-pro-oc-10g-gddr6x.html',
             'https://www.x-kom.pl/p/590450-karta-graficzna-nvidia-palit-geforce-rtx-3080-gaming-pro-oc-10g-gddr6x.html',
             'https://www.x-kom.pl/p/589758-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-eagle-oc-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/614623-karta-graficzna-nvidia-gainward-geforce-rtx-3080-phantom-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/605862-karta-graficzna-nvidia-zotac-geforce-rtx-3080-amp-holo-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/592979-karta-graficzna-nvidia-inno3d-geforce-rtx-3080-ichill-x3-new-10gb-gddr6.html',
             'https://www.x-kom.pl/p/592548-karta-graficzna-nvidia-asus-geforce-rtx-3080-rog-strix-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/591879-karta-graficzna-nvidia-zotac-geforce-rtx-3080-gaming-trinity-oc-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/590457-karta-graficzna-nvidia-gainward-geforce-rtx-3080-phoenix-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/590456-karta-graficzna-nvidia-gainward-geforce-rtx-3080-phoenix-gs-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/590452-karta-graficzna-nvidia-palit-geforce-rtx-3080-gaming-pro-10g-gddr6x.html',
             'https://www.x-kom.pl/p/590075-karta-graficzna-nvidia-asus-geforce-rtx-3080-tuf-gaming-10gb-gddr6x.html',
             'https://www.x-kom.pl/p/589762-karta-graficzna-nvidia-zotac-geforce-rtx-3080-gaming-trinity-10gb-gddr6x.html',
             '',
             '',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/600903-karta-graficzna-nvidia-msi-geforce-rtx-3090-suprim-x-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/592537-karta-graficzna-nvidia-asus-geforce-rtx-3090-rog-strix-oc-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/607812-karta-graficzna-nvidia-palit-geforce-rtx-3090-gamerock-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/605244-karta-graficzna-nvidia-gigabyte-geforce-rtx-3090-turbo-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/589754-karta-graficzna-nvidia-gigabyte-geforce-rtx-3090-gaming-oc-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/590076-karta-graficzna-nvidia-asus-geforce-rtx-3090-tuf-gaming-oc-edition-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/615653-karta-graficzna-nvidia-gigabyte-geforce-rtx-3090-aorus-wf-wb-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/615649-karta-graficzna-nvidia-gigabyte-geforce-rtx-3090-aorus-wf-xtreme-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/606125-karta-graficzna-nvidia-gigabyte-geforce-rtx-3090-aorus-xtreme-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/602253-karta-graficzna-nvidia-gigabyte-geforce-rtx-3090-master-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/592540-karta-graficzna-nvidia-asus-geforce-rtx-3090-rog-strix-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/622716-karta-graficzna-nvidia-asus-geforce-rtx-3090-ekwb-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/590448-karta-graficzna-nvidia-palit-geforce-rtx-3090-gaming-pro-oc-24g-gddr6x.html',
             'https://www.x-kom.pl/p/628032-karta-graficzna-nvidia-asus-geforce-rtx-3090-rog-strix-oc-white-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/622047-karta-graficzna-nvidia-inno3d-geforce-rtx-3090-ichill-x4-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/622045-karta-graficzna-nvidia-inno3d-geforce-rtx-3090-ichill-x3-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/622044-karta-graficzna-nvidia-inno3d-geforce-rtx-3090-gaming-x3-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/622038-karta-graficzna-nvidia-palit-geforce-rtx-3090-gamerock-oc-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/590449-karta-graficzna-nvidia-palit-geforce-rtx-3090-gaming-pro-24g-gddr6x.html',
             'https://www.x-kom.pl/p/626061-karta-graficzna-nvidia-kfa2-geforce-rtx-3090-hof-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/622080-zewnetrzna-karta-graficzna-gigabyte-geforce-rtx-3090-gaming-box-wf-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/605247-karta-graficzna-nvidia-gigabyte-geforce-rtx-3090-eagle-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/590453-karta-graficzna-nvidia-gainward-geforce-rtx-3090-phoenix-gs-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/590078-karta-graficzna-nvidia-asus-geforce-rtx-3090-tuf-gaming-24gb-gddr6x.html',
             'https://www.x-kom.pl/p/589755-karta-graficzna-nvidia-gigabyte-geforce-rtx-3090-eagle-oc-24gb-gddr6x.html',
             '',
             '',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/624270-karta-graficzna-amd-asrock-radeon-rx-6900-xt-phantom-gaming-d-16gb-gddr6.html',
             'https://www.x-kom.pl/p/631034-karta-graficzna-amd-sapphire-radeon-rx-6900-xt-nitro-16gb-gddr6.html',
             'https://www.x-kom.pl/p/624936-karta-graficzna-amd-gigabyte-radeon-rx-6900-xt-gaming-oc-16gb-gddr6.html',
             'https://www.x-kom.pl/p/624261-karta-graficzna-amd-xfx-radeon-rx-6900-xt-speedster-merc-319-16gb-gddr6.html',
             'https://www.x-kom.pl/p/615294-karta-graficzna-amd-asus-radeon-rx-6900-xt-16gb-gddr6.html',
             'https://www.x-kom.pl/p/614634-karta-graficzna-amd-asrock-radeon-rx-6900-xt-16gb-gddr6.html',
             'https://www.x-kom.pl/p/612823-karta-graficzna-amd-gigabyte-radeon-rx-6900-xt-16gb-gddr6.html',
             'https://www.x-kom.pl/p/612564-karta-graficzna-amd-msi-radeon-rx-6900-xt-16gb-gddr6.html',
             'https://www.x-kom.pl/p/612495-karta-graficzna-amd-sapphire-radeon-rx-6900-xt-16gb-gddr6.html',
             'https://www.x-kom.pl/p/625231-karta-graficzna-amd-msi-radeon-rx-6900-xt-gaming-x-trio-16gb-gddr6.html',
             'https://www.x-kom.pl/p/649516-karta-graficzna-amd-asrock-radeon-rx-6900-xt-formula-oc-16gb-gddr6.html',
             'https://www.x-kom.pl/p/656462-karta-graficzna-amd-msi-radeon-rx-6900-xt-gaming-z-trio-16gb-gddr6.html',
             '',
             '',
             'https://www.x-kom.pl/p/631032-karta-graficzna-amd-sapphire-radeon-rx-6800-gaming-oc-16gb-gddr6.html',
             'https://www.x-kom.pl/p/631031-karta-graficzna-amd-sapphire-radeon-rx-6800-xt-se-nitro-16gb-gddr6.html',
             'https://www.x-kom.pl/p/626939-karta-graficzna-amd-gigabyte-radeon-rx-6800-xt-aorus-master-type-c-16gb-gddr6.html',
             'https://www.x-kom.pl/p/609135-karta-graficzna-amd-asus-radeon-rx-6800-xt-tuf-gaming-oc-16gb-gddr6.html',
             'https://www.x-kom.pl/p/607643-karta-graficzna-amd-asrock-radeon-rx-6800-xt-taichi-x-oc-16gb-gddr6.html',
             'https://www.x-kom.pl/p/608241-karta-graficzna-amd-msi-radeon-rx-6800-gaming-x-trio-16gb-gddr6.html',
             'https://www.x-kom.pl/p/609614-karta-graficzna-amd-gigabyte-radeon-rx-6800-gaming-oc-16gb-gddr6.html',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/514368-karta-graficzna-amd-gigabyte-radeon-rx-5700-xt-gaming-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/532736-karta-graficzna-amd-gigabyte-radeon-rx-5700-xt-aorus-8g-gddr6.html',
             '',
             '',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/603491-karta-graficzna-amd-gigabyte-radeon-rx-5600-xt-gaming-oc-6gb-gddr6-rev20.html',
             'https://www.x-kom.pl/p/595152-karta-graficzna-amd-gigabyte-radeon-rx-5600-xt-windforce-oc-6g-gddr6-rev-20.html',
             'https://www.x-kom.pl/p/568064-karta-graficzna-amd-xfx-radeon-rx-5600-xt-thicc-iii-pro-6gb-gddr6.html',
             'https://www.x-kom.pl/p/541026-karta-graficzna-amd-xfx-radeon-rx-5600-xt-thicc-iii-ultra-6gb-gddr6.html',
             'https://www.x-kom.pl/p/541022-karta-graficzna-amd-powercolor-radeon-rx-5600-xt-red-devil-6gb-gddr6.html',
             'https://www.x-kom.pl/p/540866-karta-graficzna-amd-gigabyte-radeon-rx-5600-xt-windforce-oc-6gb-gddr6.html',
             'https://www.x-kom.pl/p/539800-karta-graficzna-amd-msi-radeon-rx-5600-xt-gaming-x-6gb-gddr6.html',
             'https://www.x-kom.pl/p/538460-karta-graficzna-amd-asrock-radeon-rx-5600-xt-challenger-d-oc-6gb-gddr6.html',
             'https://www.x-kom.pl/p/538459-karta-graficzna-amd-asrock-radeon-rx-5600-xt-phantom-gaming-d2-oc-6gb-gddr6.html',
             'https://www.x-kom.pl/p/538453-karta-graficzna-amd-asrock-radeon-rx-5600-xt-phantom-gaming-d3-oc-6gb-gddr6.html',
             '',
             '',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/533867-karta-graficzna-amd-sapphire-radeon-rx-5500-xt-pulse-4gb-gddr6.html',
             'https://www.x-kom.pl/p/625906-karta-graficzna-amd-amd-radeon-pro-w5500-8gb-gddr6.html',
             'https://www.x-kom.pl/p/533896-karta-graficzna-amd-gigabyte-radeon-rx-5500-xt-oc-8gb-gddr6.html',
             'https://www.x-kom.pl/p/602639-karta-graficzna-amd-gigabyte-radeon-rx-5500-xt-oc-8gb-gddr6-rev20.html',
             'https://www.x-kom.pl/p/540912-karta-graficzna-amd-msi-radeon-rx-5500-xt-mech-oc-4gb-gddr6.html',
             '',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/630290-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-gaming-oc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/624768-karta-graficzna-nvidia-asus-geforce-rtx-3060-tuf-gaming-oc-edition-12gb-gddr6.html',
             'https://www.x-kom.pl/p/630288-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-eagle-oc-12gd-gddr6.html',
             'https://www.x-kom.pl/p/630711-karta-graficzna-nvidia-msi-geforce-rtx-3060-gaming-x-12gb-gddr6.html',
             'https://www.x-kom.pl/p/633020-karta-graficzna-nvidia-zotac-geforce-rtx-3060-twin-edge-oc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/632067-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-vision-oc-12g-12gb-gddr6.html',
             'https://www.x-kom.pl/p/633004-karta-graficzna-nvidia-inno3d-geforce-rtx-3060-twin-x2-oc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/624769-karta-graficzna-nvidia-asus-geforce-rtx-3060-rog-strix-oc-edition-12gb-gddr6.html',
             'https://www.x-kom.pl/p/630713-karta-graficzna-nvidia-msi-geforce-rtx-3060-gaming-x-trio-12gb-gddr6.html',
             'https://www.x-kom.pl/p/632069-karta-graficzna-nvidia-gigabyte-geforce-rtx-3060-eagle-12gb-gddr6.html',
             'https://www.x-kom.pl/p/630712-karta-graficzna-nvidia-msi-geforce-rtx-3060-ventus-2x-oc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/633021-karta-graficzna-nvidia-zotac-geforce-rtx-3060-twin-edge-12gb-gddr6.html',
             'https://www.x-kom.pl/p/632856-karta-graficzna-nvidia-msi-geforce-rtx-3060-ventus-3x-oc-12gb.html',
             'https://www.x-kom.pl/p/633455-karta-graficzna-nvidia-palit-geforce-rtx-3060-dual-12gb-gddr6.html',
             'https://www.x-kom.pl/p/633459-karta-graficzna-nvidia-palit-geforce-rtx-3060-dual-oc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/633463-karta-graficzna-nvidia-gainward-geforce-rtx-3060-pegasus-12gb-gddr6.html',
             'https://www.x-kom.pl/p/633464-karta-graficzna-nvidia-gainward-geforce-rtx-3060-pegasus-oc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/633465-karta-graficzna-nvidia-gainward-geforce-rtx-3060-ghost-12gb-gddr6.html',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/638180-karta-graficzna-amd-asus-radeon-rx-6700-xt-tuf-gaming-oc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/638549-karta-graficzna-amd-msi-radeon-rx-6700-xt-mech-2x-12gb-gddr6.html',
             'https://www.x-kom.pl/p/638550-karta-graficzna-amd-msi-radeon-rx-6700-xt-mech-2x-oc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/638551-karta-graficzna-amd-msi-radeon-rx-6700-xt-gaming-x-12gb-gddr6.html',
             'https://www.x-kom.pl/p/638552-karta-graficzna-amd-msi-radeon-rx-6700-xt-12gb-gddr6.html',
             'https://www.x-kom.pl/p/640049-karta-graficzna-amd-gigabyte-radeon-rx-6700-xt-eagle-12gb-gddr6.html',
             'https://www.x-kom.pl/p/640189-karta-graficzna-amd-asrock-radeon-rx-6700-xt-12gb-gddr6.html',
             'https://www.x-kom.pl/p/640756-karta-graficzna-amd-sapphire-radeon-rx-6700-xt-12gb-gddr6.html',
             'https://www.x-kom.pl/p/640759-karta-graficzna-amd-xfx-radeon-rx-6700-xt-gaming-speedster-12gb-gddr6.html',
             'https://www.x-kom.pl/p/640761-karta-graficzna-amd-xfx-radeon-rx-6700-xt-ultra-speedster-qick-12gb-gddr6.html',
             'https://www.x-kom.pl/p/640762-karta-graficzna-amd-xfx-radeon-rx-6700-xt-speedster-merc-12gb-gddr6.html',
             'https://www.x-kom.pl/p/641503-karta-graficzna-amd-sapphire-radeon-rx-6700-xt-pulse-12gb-gddr6.html',
             'https://www.x-kom.pl/p/641504-karta-graficzna-amd-sapphire-radeon-rx-6700-xt-nitro-12gb-gddr6.html',
             'https://www.x-kom.pl/p/641839-karta-graficzna-amd-asus-radeon-rx-6700-xt-12gb-gddr6.html',
             'https://www.x-kom.pl/p/637730-karta-graficzna-amd-gigabyte-radeon-rx-6700-xt-gaming-oc-12gb-gddr6.html',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/658146-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-ti-gaming-oc-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/658152-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-ti-eagle-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/657483-karta-graficzna-nvidia-asus-geforce-rtx3080-ti-rog-strix-gaming-oc-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/655240-karta-graficzna-nvidia-msi-geforce-rtx-3080-ti-suprim-x-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/658574-karta-graficzna-nvidia-asus-geforce-rtx-3080-ti-tuf-gaming-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/658452-karta-graficzna-nvidia-palit-geforce-rtx-3080-ti-gamerock-oc-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/658069-karta-graficzna-nvidia-palit-geforce-rtx-3080-ti-gamingpro-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/657482-karta-graficzna-nvidia-asus-geforce-rtx-3080-ti-tuf-gaming-oc-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/657425-karta-graficzna-nvidia-kfa2-geforce-rtx-3080-ti-sg-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/659626-karta-graficzna-nvidia-evga-geforce-rtx-3080-ti-ftw3-ultra-gaming-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/658147-karta-graficzna-nvidia-gigabyte-geforce-rtx-3080-ti-vision-oc-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/658068-karta-graficzna-nvidia-gainward-geforce-rtx-3080-ti-phoenix-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/658067-karta-graficzna-nvidia-palit-geforce-rtx-3080-ti-gamerock-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/656410-karta-graficzna-nvidia-inno3d-geforce-rtx-3080-ti-ichill-x4-12gb-gddr6x.html',
             'https://www.x-kom.pl/p/655242-karta-graficzna-nvidia-msi-geforce-rtx-3080-ti-gaming-x-trio-12gb-gddr6x.html',
             '',
             '',
             '',
             'https://www.x-kom.pl/p/658888-karta-graficzna-nvidia-inno3d-geforce-rtx-3070-ti-x3-oc-8gb-gddr6x.html',
             'https://www.x-kom.pl/p/658547-karta-graficzna-nvidia-palit-geforce-rtx-3070-ti-gaming-pro-8gb-gddr6x.html',
             'https://www.x-kom.pl/p/655245-karta-graficzna-nvidia-msi-geforce-rtx-3070-ti-ventus-3x-oc-8gb-gddr6x.html',
             'https://www.x-kom.pl/p/655243-karta-graficzna-nvidia-msi-geforce-rtx-3070-ti-suprim-x-8gb-gddr6x.html',
             'https://www.x-kom.pl/p/658552-karta-graficzna-nvidia-palit-geforce-rtx-3070-ti-game-rock-oc-8gb-gddr6x.html',
             '',
             '',
             '',
             '',
             '',
             '']
    i = 0
    for x in linki:
        try:
            if x != '':
                print(x)
                req = Request(x)
                response = urllib.request.urlopen(req)
                html = response.read()
                if response.getcode() != 200:
                    print('continue')
                    continue
                soup = BeautifulSoup(html, 'html.parser')
                try:
                    slownik[i, 0] = soup.find('h1', {'class': ['sc-1bker4h-4 driGYx']}).text  # nazwa
                except:
                    slownik[i, 0] = soup.find('h1', {'class': ['sc-1bker4h-4 fiaogA']}).text  # nazwa
                slownik[i, 1] = soup.find('div', {'class': ['u7xnnm-4 jFbqvs']}).text  # aktualnaCena
                slownik[i, 1] = slownik[i, 1][:8]  # formatowanie ceny
                slownik[i, 1] = slownik[i, 1].replace(" z", "")
                slownik[i, 1] = slownik[i, 1].replace(" ", "")
                slownik[i, 2] = soup.find('div', {'class': ['u7xnnm-3 cvqevx']})  # poprzedniaCena
                if slownik[i, 2]:
                    slownik[i, 2] = True
                else:
                    slownik[i, 2] = False
                slownik[i, 3] = soup.find('span',
                                          {'class': ['sc-1smss4h-5 iZjlqx']})  # dostepnosc
                slownik[i, 3] = str(slownik[i, 3])
                if 'inne produkty' in slownik[i, 3] or 'powiadom' in slownik[i, 3] or 'dostępnoś' in slownik[i, 3]:
                    slownik[i, 3] = False
                else:
                    slownik[i, 3] = True
                slownik[i, 4] = x  # link
            else:
                slownik[i, 0] = ''
                slownik[i, 1] = ''
                slownik[i, 2] = ''
                slownik[i, 3] = ''
                slownik[i, 4] = ''
            i = i + 1
        except:
            i = i + 1
    return i - 1


def zapisDatyXkom(p):
    plik = openpyxl.load_workbook(p)
    odl = 0
    for x in range(5, 1000):
        if plik["Arkusz1"].cell(row=1, column=x).value == aktualnaData and plik["Arkusz1"].cell(row=2,
                                                                                                column=x).value != None:
            print(plik["Arkusz1"].cell(row=2, column=x).value)
            return odl
        elif plik["Arkusz1"].cell(row=1, column=x).value is None or (
                plik["Arkusz1"].cell(row=1, column=x).value == aktualnaData and plik["Arkusz1"].cell(row=2,
                                                                                                     column=x).value == None):
            plik["Arkusz1"].cell(row=1, column=x).value = aktualnaData
            plik["Arkusz1"].cell(row=1, column=x).fill = PatternFill(fgColor="C6E0B4", fill_type="solid")
            odl = x
            break
    plik.save(p)
    plik.close()
    return odl


def zapisDanychXkom(odl, p):
    plik = openpyxl.load_workbook(p)
    i = 0
    ile = 0
    for x in slownik:
        # print(f'{x}  {ile}  {i}')
        # print(ile)
        if ile == 0:
            plik["Arkusz1"].cell(row=i + 2, column=2).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=2).value = slownik[i, 0]
            ile += 1
        elif ile == 1:
            if str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) > slownik[i, 1] and str(
                    plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) != None and slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="E6B8B7", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) > slownik[i, 1] and str(
                    plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) != None:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="FFE699", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) < slownik[i, 1] and slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="632523", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) < slownik[i, 1]:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="F4B084", fill_type="solid")
            elif slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=odl).value = slownik[i, 1]
            ile += 1
        elif ile == 2:
            if slownik[i, 2] == True:
                plik["Arkusz1"].cell(row=i + 2, column=4).fill = PatternFill(fgColor="00B050", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=4).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=4).value = slownik[i, 2]
            ile += 1
        elif ile == 3:
            if slownik[i, 3] == True:
                plik["Arkusz1"].cell(row=i + 2, column=3).fill = PatternFill(fgColor="00B050", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=3).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=3).value = slownik[i, 3]
            ile += 1

        else:
            plik["Arkusz1"].cell(row=i + 2, column=1).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=1).value = slownik[i, 4]
            ile = 0
            i += 1

    plik.save(p)
    plik.close()


def pobieranieRTV():
    linki = ['https://www.euro.com.pl/procesory/amd-procesor-amd-ryzen-5-3600.bhtml',
             '',
             'https://www.euro.com.pl/karty-graficzne/sapphire-karta-graf-sapphire-pulse-rx5700xt-8gb.bhtml',
             '',
             'https://www.euro.com.pl/plyty-glowne/asus-plyta-glowna-asus-tuf-b450-pro-gaming.bhtml',
             'https://www.euro.com.pl/plyty-glowne/msi-plyta-glowna-msi-mpg-x570-gam-edge-wifi_1.bhtml',
             '',
             'https://www.euro.com.pl/dyski-wewnetrzne-ssd/adata-dysk-adata-ssd-xpg-s11-512gb-pcie-m-2.bhtml',
             '',
             '',
             'https://www.euro.com.pl/pamieci-ram/goodram-pamiecpc-good-ddr4irdmpro16-36002x8cza.bhtml',
             '',
             '',
             'https://www.euro.com.pl/zasilacze-do-komputerow-pc/silentiumpc-supremo-m2-gold-550w.bhtml',
             'https://www.euro.com.pl/zasilacze-do-komputerow-pc/silentiumpc-supremo-fm2-gold-750w-80-gold.bhtml',
             '',
             'https://www.euro.com.pl/chlodzenie-procesory/silentiumpc-chlodzenie-cpu-sil-pc-fortis-3-evo-argb.bhtml',
             '',
             '',
             '',
             'https://www.euro.com.pl/obudowy-pc/silentiumpc-obudowa-pc-sile-pcsignum-sg1qevo-tgargb.bhtml',
             'https://www.euro.com.pl/obudowy-pc/silentiumpc-obudowapc-sile-pc-regnumrg6vtgpure-black.bhtml',
             'https://www.euro.com.pl/obudowy-pc/silentiumpc-obudowa-pc-sile-reg-rg6vevotgargbpurebla.bhtml',
             '',
             '',
             'https://www.euro.com.pl/monitory-led-i-lcd/aoc-agon-ag241qx.bhtml',
             '',
             'https://www.euro.com.pl/chlodzenie-procesory/silentiumpc-chlodzenie-sil-grandis-3.bhtml',
             'https://www.euro.com.pl/plyty-glowne/asus-plyta-glowna-asus-rogstrixb550fgamwifi.bhtml']
    i = 0
    for x in linki:
        try:
            if x != '':
                print(x)
                req = Request(x, headers={"User-Agent": "Mozilla/5.0"})
                response = urllib.request.urlopen(req)
                html = response.read()
                if response.getcode() != 200:
                    print('continue')
                    continue
                soup = BeautifulSoup(html, 'html.parser')
                try:
                    slownik[i, 0] = soup.find('h1', {'class': ['product-name selenium-KP-product-name']}).text  # nazwa
                except:
                    slownik[i, 0] = soup.find('title').text  # nazwa2
                try:
                    slownik[i, 1] = soup.find('div',
                                              {'class': ['product-price selenium-price-normal']}).text  # aktualnaCena
                except:
                    slownik[i, 1] = ''
                slownik[i, 1] = slownik[i, 1].replace("z", "")
                slownik[i, 1] = slownik[i, 1].replace("ł", "")
                slownik[i, 1] = slownik[i, 1].replace(" ", "")
                slownik[i, 2] = soup.find('div', {'class': ['price-old']})  # poprzedniaCena
                if slownik[i, 2]:
                    slownik[i, 2] = True
                else:
                    slownik[i, 2] = False
                slownik[i, 3] = soup.find('div',
                                          {'class': ['temporary-unavailable']})  # dostepnosc
                slownik[i, 3] = str(slownik[i, 3])
                if 'nie jest dostępny' in slownik[i, 3] or 'niedostępny' in slownik[i, 3]:
                    slownik[i, 3] = False
                else:
                    slownik[i, 3] = True
                if slownik[i, 3]:
                    slownik[i, 3] = soup.find('div',
                                              {'class': ['product-unavailable']})  # dostepnosc2
                    slownik[i, 3] = str(slownik[i, 3])
                    if 'nie jest dostępny' in slownik[i, 3] or 'niedostępny' in slownik[i, 3]:
                        slownik[i, 3] = False
                    else:
                        slownik[i, 3] = True
                slownik[i, 4] = x  # link
            else:
                slownik[i, 0] = ''
                slownik[i, 1] = ''
                slownik[i, 2] = ''
                slownik[i, 3] = ''
                slownik[i, 4] = ''
            i = i + 1
        except:
            i = i + 1
    return i - 1


def zapisDatyRTV(p):
    plik = openpyxl.load_workbook(p)
    odl = 0
    for x in range(5, 1000):
        if plik["Arkusz1"].cell(row=1, column=x).value == aktualnaData and plik["Arkusz1"].cell(row=2,
                                                                                                column=x).value != None:
            print(plik["Arkusz1"].cell(row=2, column=x).value)
            return odl
        elif plik["Arkusz1"].cell(row=1, column=x).value is None or (
                plik["Arkusz1"].cell(row=1, column=x).value == aktualnaData and plik["Arkusz1"].cell(row=2,
                                                                                                     column=x).value == None):
            plik["Arkusz1"].cell(row=1, column=x).value = aktualnaData
            plik["Arkusz1"].cell(row=1, column=x).fill = PatternFill(fgColor="C6E0B4", fill_type="solid")
            odl = x
            break
    plik.save(p)
    plik.close()
    return odl


def zapisDanychRTV(odl, p):
    plik = openpyxl.load_workbook(p)
    i = 0
    ile = 0
    for x in slownik:
        # print(f'{x}  {ile}  {i}')
        # print(ile)
        if ile == 0:
            plik["Arkusz1"].cell(row=i + 2, column=2).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=2).value = slownik[i, 0]
            ile += 1
        elif ile == 1:
            if str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) > slownik[i, 1] and str(
                    plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) != None and slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="E6B8B7", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) > slownik[i, 1] and str(
                    plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) != None:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="FFE699", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) < slownik[i, 1] and slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="632523", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) < slownik[i, 1]:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="F4B084", fill_type="solid")
            elif slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=odl).value = slownik[i, 1]
            ile += 1
        elif ile == 2:
            if slownik[i, 2]:
                plik["Arkusz1"].cell(row=i + 2, column=4).fill = PatternFill(fgColor="00B050", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=4).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=4).value = slownik[i, 2]
            ile += 1
        elif ile == 3:
            if slownik[i, 3]:
                plik["Arkusz1"].cell(row=i + 2, column=3).fill = PatternFill(fgColor="00B050", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=3).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=3).value = slownik[i, 3]
            ile += 1

        else:
            plik["Arkusz1"].cell(row=i + 2, column=1).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=1).value = slownik[i, 4]
            ile = 0
            i += 1

    plik.save(p)
    plik.close()


def pobieranieMedia():
    linki = [
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/procesory/procesor-amd-ryzen-5-3600',
        '',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/karty-graficzne/sapphire-pulse-radeon-rx-5700-xt-8g-gddr6-hdmi-triple-dp-oc-w-bp-uefi-11293-01-20g',
        '',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/plyty-glowne/asus-tuf-b450-pro-gaming-am4-b450-ddr4-3533mhz-dual-m-2-dvi-d-hdmi-tuf-b450-pro-gaming',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/plyty-glowne/plyta-glowna-msi-mpg-x570-gaming-edge-wi-fi',
        '',
        '',
        '',
        '',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/pamieci-ram/goodram-irdm-pro-pamiec-ddr4-16gb-3600mhz-cl17-1-35v-czarna-irp-3600d4v64l17-16g',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/pamieci-ram/pamiec-ram-crucial-ballistix-16gb-3600mhz-ddr4-cl16-dimm-2x8-white-rgb',
        '',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/zasilacze/supremo-m2-550w-80-gold-psu-modular',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/zasilacze/supremo-fm2-gold-650w-modular',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/chlodzenie/chlodzenie-cpu-fortis-3-he1425',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/chlodzenie/wentylator-silentiumpc-fortis-3-evo-argb-he1425',
        '',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/obudowy/obudowa-komputerowa-silentium-pc-armis-ar7x-evo-tg-argb',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/obudowy/obudowa-komputerowa-silentium-pc-armis-ar6q-evo-tg-argb',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/obudowy/obudowa-pc-signum-sg1q-evo-tg-argb',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/obudowy/obudowa-komputerowa-silentium-pc-regnum-rg6v-tg',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/obudowy/obudowa-komputerowa-silentium-pc-regnum-rg6v-evo-argb',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/obudowy/obudowa-komputerowa-msi-mag-vampiric-100',
        '',
        'https://www.mediaexpert.pl/komputery-i-tablety/monitory-led/monitor-aoc-ag241qx',
        '',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/chlodzenie/chlodzenie-cpu-silentium-pc-grandis-3',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/plyty-glowne/asus-rog-strix-b550-f-gaming-wi-fi-am4-ddr4-2xm-2-6xsata-usb3-2-atx-mb-rog-strix-b550-f-gaming-wi-fi',
        'https://www.mediaexpert.pl/komputery-i-tablety/dyski-i-pamieci/dyski-wewnetrzne/dysk-ssd-patriot-p210-1tb-2-5-sata-iii',
        'https://www.mediaexpert.pl/komputery-i-tablety/podzespoly-komputerowe/pamieci-ram/patriot-viper-steel-series-ddr4-2x8gb-3600mhz-xmp2']
    i = 0
    for x in linki:
        try:
            if x != '':
                print(x)
                req = Request(x, headers={"User-Agent": "Mozilla/5.0"})
                response = urllib.request.urlopen(req)
                html = response.read()
                if response.getcode() != 200:
                    print('continue')
                    continue
                soup = BeautifulSoup(html, 'html.parser')
                slownik[i, 0] = soup.find('h1', {'class': ['a-typo is-primary']}).text  # nazwa
                try:
                    slownik[i, 1] = soup.find('span', {'class': ['a-price_price']}).text  # aktualnaCena
                except:
                    slownik[i, 1] = ''
                slownik[i, 1] = slownik[i, 1].replace("z", "")
                slownik[i, 1] = slownik[i, 1].replace("ł", "")
                slownik[i, 1] = slownik[i, 1].replace(" ", "")
                slownik[i, 2] = soup.find('div',
                                          {'class': ['c-offerBox_discount ']})  # poprzedniacena
                slownik[i, 2] = str(slownik[i, 2])
                if 'Taniej o' in slownik[i, 2]:
                    slownik[i, 2] = True
                else:
                    slownik[i, 2] = False
                if slownik[i, 2] is False:
                    slownik[i, 2] = soup.find('p',
                                              {'class': ['is-firstRow']})  # poprzedniacena2
                    slownik[i, 2] = str(slownik[i, 2])
                    if 'Cena z kodem' in slownik[i, 2] or 'NOCNA' in slownik[i, 2]:
                        slownik[i, 2] = True
                    else:
                        slownik[i, 2] = False
                slownik[i, 3] = soup.find('div',
                                          {'class': ['a-typo is-tertiary']})  # dostepnosc
                slownik[i, 3] = str(slownik[i, 3])
                if 'w wybranych sklepach' in slownik[i, 3] or 'Produkt' in slownik[i, 3]:
                    slownik[i, 3] = False
                else:
                    slownik[i, 3] = True
                slownik[i, 4] = x  # link
            else:
                slownik[i, 0] = ''
                slownik[i, 1] = ''
                slownik[i, 2] = ''
                slownik[i, 3] = ''
                slownik[i, 4] = ''
            i = i + 1
        except:
            i = i + 1
    return i - 1


def zapisDatyMedia(p):
    plik = openpyxl.load_workbook(p)
    odl = 0
    for x in range(5, 1000):
        if plik["Arkusz1"].cell(row=1, column=x).value == aktualnaData and plik["Arkusz1"].cell(row=2,
                                                                                                column=x).value != None:
            print(plik["Arkusz1"].cell(row=2, column=x).value)
            return odl
        elif plik["Arkusz1"].cell(row=1, column=x).value is None or (
                plik["Arkusz1"].cell(row=1, column=x).value == aktualnaData and plik["Arkusz1"].cell(row=2,
                                                                                                     column=x).value == None):
            plik["Arkusz1"].cell(row=1, column=x).value = aktualnaData
            plik["Arkusz1"].cell(row=1, column=x).fill = PatternFill(fgColor="C6E0B4", fill_type="solid")
            odl = x
            break
    plik.save(p)
    plik.close()
    return odl


def zapisDanychMedia(odl, p):
    plik = openpyxl.load_workbook(p)
    i = 0
    ile = 0
    for x in slownik:
        # print(f'{x}  {ile}  {i}')
        # print(ile)
        if ile == 0:
            plik["Arkusz1"].cell(row=i + 2, column=2).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=2).value = slownik[i, 0]
            ile += 1
        elif ile == 1:
            if str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) > slownik[i, 1] and str(
                    plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) != None and slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="E6B8B7", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) > slownik[i, 1] and str(
                    plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) != None:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="FFE699", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) < slownik[i, 1] and slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="632523", fill_type="solid")
            elif str(plik["Arkusz1"].cell(row=i + 2, column=odl - 1).value) < slownik[i, 1]:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="F4B084", fill_type="solid")
            elif slownik[i, 3] == False:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=odl).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=odl).value = slownik[i, 1]
            ile += 1
        elif ile == 2:
            if slownik[i, 2]:
                plik["Arkusz1"].cell(row=i + 2, column=4).fill = PatternFill(fgColor="00B050", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=4).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=4).value = slownik[i, 2]
            ile += 1
        elif ile == 3:
            if slownik[i, 3]:
                plik["Arkusz1"].cell(row=i + 2, column=3).fill = PatternFill(fgColor="00B050", fill_type="solid")
            else:
                plik["Arkusz1"].cell(row=i + 2, column=3).fill = PatternFill(fgColor="FF0000", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=3).value = slownik[i, 3]
            ile += 1

        else:
            plik["Arkusz1"].cell(row=i + 2, column=1).fill = PatternFill(fgColor="E2EFDA", fill_type="solid")
            plik["Arkusz1"].cell(row=i + 2, column=1).value = slownik[i, 4]
            ile = 0
            i += 1

    plik.save(p)
    plik.close()

def xkom():
    try:
        pobieranieXkom()
    except:
        pass
    try:
        path1 = 'C:\\Users\\piotr\\OneDrive - T-Mobile Polska S.A\\xkom.xlsx'
        path2 = 'C:\\Users\\pkaniewski3\\OneDrive - T-Mobile Polska S.A\\xkom.xlsx'
        try:
            odl = zapisDatyXkom(path1)
            if odl != 0:
                zapisDanychXkom(odl, path1)
                print(path1)
                print('zapis xKom1')
        except:
            odl = zapisDatyXkom(path2)
            if odl != 0:
                zapisDanychXkom(odl, path2)
                print(path2)
                print('zapis xKom2')
        slownik.clear()
    except Exception as e:
        print(e.args)
        print(e.message)
        pass

def morele():
    try:
        pobieranieMorele()
    except:
        pass
    try:
        path1 = 'C:\\Users\\piotr\\OneDrive - T-Mobile Polska S.A\\morele.xlsx'
        path2 = 'C:\\Users\\pkaniewski3\\OneDrive - T-Mobile Polska S.A\\morele.xlsx'
        try:
            odl = zapisDatyMorele(path1)
            if odl != 0:
                zapisDanychMorele(odl, path1)
                print('zapis Morele1')
        except:
            odl = zapisDatyMorele(path2)
            if odl != 0:
                zapisDanychMorele(odl, path2)
                print('zapis Morele2')
        slownik.clear()
    except Exception as e:
        print(e.args)
        print(e.message)
        pass
def rtv():
    try:
        pobieranieRTV()
    except:
        pass
    try:
        path1 = 'C:\\Users\\piotr\\OneDrive - T-Mobile Polska S.A\\rtv.xlsx'
        path2 = 'C:\\Users\\pkaniewski3\\OneDrive - T-Mobile Polska S.A\\rtv.xlsx'
        try:
            odl = zapisDatyRTV(path1)
            if odl != 0:
                zapisDanychRTV(odl, path1)
                print('zapis RTV1')
        except:
            odl = zapisDatyRTV(path2)
            if odl != 0:
                zapisDanychRTV(odl, path2)
                print('zapis RTV2')
        slownik.clear()
    except Exception as e:
        print(e.args)
        print(e.message)
        pass

def media():
    try:
            pobieranieMedia()
    except:
        pass
    try:
        path1 = 'C:\\Users\\piotr\\OneDrive - T-Mobile Polska S.A\\mediaexpert.xlsx'
        path2 = 'C:\\Users\\pkaniewski3\\OneDrive - T-Mobile Polska S.A\\mediaexpert.xlsx'
        try:
            odl = zapisDatyMedia(path1)
            if odl != 0:
                zapisDanychMedia(odl, path1)
                print('zapis Media1')
        except:
            odl = zapisDatyMedia(path2)
            if odl != 0:
                zapisDanychMedia(odl, path2)
                print('zapis Media2')
    except Exception as e:
        print(e.args)
        print(e.message)
        pass


def main():
    morele()
    xkom()
    rtv()  
    media()
    
    
    


if __name__ == "__main__":
    main()
